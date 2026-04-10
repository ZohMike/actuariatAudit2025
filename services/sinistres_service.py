"""
Service de gestion des sinistres
- Chargement des fichiers sinistres par année
- Normalisation des colonnes (différentes selon l'année)
- Calcul des règlements et SAP
"""
import streamlit as st
import polars as pl
import pandas as pd
import io
import re
from typing import Optional

from utils.excel_io import read_excel_to_polars


class SinistresService:
    """Service pour le traitement des données sinistres."""
    
    @staticmethod
    def normalize_col_name(name: str) -> str:
        """Normalise un nom de colonne pour comparaison."""
        # Supprimer accents
        name = name.lower().strip()
        name = name.replace("û", "u").replace("é", "e").replace("è", "e").replace("ê", "e")
        # Supprimer espaces multiples et caractères spéciaux
        name = re.sub(r'\s+', ' ', name).strip()  # Garder un seul espace
        return name
    
    @staticmethod
    def find_column(df: pl.DataFrame, exact_names: list[str], partial_patterns: list[str] = None) -> Optional[str]:
        """
        Trouve le nom de colonne réel.
        
        Args:
            df: DataFrame
            exact_names: Noms exacts à chercher (priorité)
            partial_patterns: Patterns partiels (fallback)
            
        Returns:
            Nom réel de la colonne ou None
        """
        # D'abord chercher correspondance exacte (après normalisation)
        for col in df.columns:
            normalized = SinistresService.normalize_col_name(col)
            for exact in exact_names:
                if normalized == exact.lower():
                    return col
        
        # Ensuite chercher patterns partiels si fournis
        if partial_patterns:
            for col in df.columns:
                normalized = SinistresService.normalize_col_name(col)
                # Supprimer espaces pour comparaison partielle
                normalized_no_space = normalized.replace(" ", "")
                for pattern in partial_patterns:
                    if pattern in normalized_no_space:
                        return col
        
        return None
    
    @staticmethod
    def _amount_expr(col_name: str) -> pl.Expr:
        """
        Convertit montants Excel (tirets, '- ', vides, virgule décimale, espaces milliers) en Float64.
        """
        s = pl.col(col_name).cast(pl.Utf8, strict=False).str.strip_chars()
        low = s.str.to_lowercase()
        s = (
            pl.when(s.is_null() | (s.str.len_chars() == 0))
            .then(None)
            .when(
                low.is_in(
                    [
                        "-",
                        "—",
                        "–",
                        "--",
                        "---",
                        "- ",
                        " -",
                        "n/a",
                        "na",
                        "#n/a",
                        "#na",
                        "nan",
                        "none",
                        ".",
                        "..",
                    ]
                )
            )
            .then(None)
            .when(s.str.contains(r"^[\s\-\u2013\u2014\u00a0]+$"))
            .then(None)
            .otherwise(s)
        )
        s = s.str.replace(",", ".").str.replace_all(r"\s+", "")
        return s.cast(pl.Float64, strict=False).fill_null(0.0)
    
    @staticmethod
    def normalize_columns(df: pl.DataFrame) -> pl.DataFrame:
        """Normalise les noms de colonnes du DataFrame sinistres."""
        
        # Colonnes attendues: BRANCHE, GARANTIE, REGLEMENT, COUT_TOTAL
        # Noms exacts à chercher (priorité), puis patterns partiels (fallback)
        branche_exact = ["branche"]
        branche_partial = None
        
        reglement_exact = ["reglement", "règlement"]
        reglement_partial = ["reglement", "règlement"]
        
        cout_exact = ["cout_total", "cout total", "coût_total", "coût total"]
        cout_partial = ["couttotal", "coutotal"]
        
        # Trouver les colonnes
        branche_col = SinistresService.find_column(df, branche_exact, branche_partial)
        reglement_col = SinistresService.find_column(df, reglement_exact, reglement_partial)
        cout_col = SinistresService.find_column(df, cout_exact, cout_partial)
        
        # Debug: afficher les colonnes trouvées
        if not all([branche_col, reglement_col, cout_col]):
            missing = []
            found_info = []
            if not branche_col:
                missing.append("BRANCHE")
            else:
                found_info.append(f"Branche='{branche_col}'")
            if not reglement_col:
                missing.append("REGLEMENT")
            else:
                found_info.append(f"Reglement='{reglement_col}'")
            if not cout_col:
                missing.append("COUT_TOTAL")
            else:
                found_info.append(f"Cout='{cout_col}'")
            
            # Afficher les colonnes disponibles pour debug
            cols_disponibles = ", ".join(df.columns)
            raise ValueError(
                f"Colonnes manquantes: {', '.join(missing)}. "
                f"Colonnes trouvées: {', '.join(found_info) if found_info else 'aucune'}. "
                f"Colonnes disponibles: {cols_disponibles}"
            )
        
        # Sélectionner et renommer (montants : parsing tolérant aux tirets / cellules vides Excel)
        return df.select([
            pl.col(branche_col).cast(pl.Utf8, strict=False).fill_null("").str.strip_chars().alias("Branche"),
            SinistresService._amount_expr(reglement_col).alias("Total_Reglement"),
            SinistresService._amount_expr(cout_col).alias("Cout_Total"),
        ])
    
    @staticmethod
    def get_unique_branches(file_contents: list[tuple[str, bytes, str]]) -> list[str]:
        """
        Extrait les branches uniques de tous les fichiers sinistres.
        
        Args:
            file_contents: Liste de tuples (nom_fichier, contenu, année)
            
        Returns:
            Liste triée des branches uniques
        """
        import openpyxl
        
        all_branches = set()
        
        for filename, content, year in file_contents:
            buffer = io.BytesIO(content)
            
            # Lire le fichier (sinistres : tout en texte puis parsing montants tolérant)
            if filename.endswith('.csv'):
                buffer.seek(0)
                try:
                    pdf = pd.read_csv(
                        buffer, dtype=str, keep_default_na=False, encoding="utf-8-sig"
                    )
                except UnicodeDecodeError:
                    buffer.seek(0)
                    pdf = pd.read_csv(
                        buffer, dtype=str, keep_default_na=False, encoding="latin-1"
                    )
                df = pl.from_pandas(pdf)
            else:
                # Pour Excel, lister les feuilles et choisir la bonne
                buffer_check = io.BytesIO(content)
                wb = openpyxl.load_workbook(buffer_check, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                
                # Chercher une feuille qui contient "sinistre" ou prendre la première
                target_sheet = None
                for sheet in sheet_names:
                    if "sinistre" in sheet.lower():
                        target_sheet = sheet
                        break
                
                if target_sheet is None:
                    target_sheet = sheet_names[0]
                
                df = read_excel_to_polars(
                    content, filename, sheet_name=target_sheet, string_cells=True
                )
            
            # Normaliser et extraire les branches
            df_normalized = SinistresService.normalize_columns(df)
            branches = df_normalized["Branche"].unique().to_list()
            all_branches.update([b for b in branches if b is not None])
        
        return sorted(all_branches)
    
    @staticmethod
    def load_sinistres_files(
        file_contents: list[tuple[str, bytes, str]],  # (filename, content, year)
        mapping: tuple
    ) -> pl.DataFrame:
        """
        Charge et traite les fichiers sinistres.
        
        Args:
            file_contents: Liste de tuples (nom_fichier, contenu, année)
            mapping: Mapping des branches (tuple pour cache)
            
        Returns:
            DataFrame avec Branche, Exercice, Total_Reglement, Cout_Total, SAP
        """
        import openpyxl
        
        mapping_dict = dict(mapping)
        all_data = []
        
        for filename, content, year in file_contents:
            buffer = io.BytesIO(content)
            
            # Lire le fichier
            sheet_info = "CSV"
            if filename.endswith('.csv'):
                buffer.seek(0)
                try:
                    pdf = pd.read_csv(
                        buffer, dtype=str, keep_default_na=False, encoding="utf-8-sig"
                    )
                except UnicodeDecodeError:
                    buffer.seek(0)
                    pdf = pd.read_csv(
                        buffer, dtype=str, keep_default_na=False, encoding="latin-1"
                    )
                df = pl.from_pandas(pdf)
            else:
                # Pour Excel, lister les feuilles et choisir la bonne
                buffer_check = io.BytesIO(content)
                wb = openpyxl.load_workbook(buffer_check, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                
                # Chercher une feuille qui contient "sinistre" ou prendre la première
                target_sheet = None
                for sheet in sheet_names:
                    if "sinistre" in sheet.lower():
                        target_sheet = sheet
                        break
                
                if target_sheet is None:
                    target_sheet = sheet_names[0]
                
                sheet_info = f"{target_sheet} (feuilles: {', '.join(sheet_names)})"
                
                df = read_excel_to_polars(
                    content, filename, sheet_name=target_sheet, string_cells=True
                )
            
            # Debug: Afficher les colonnes lues
            st.write(f"📄 **{filename}** → {sheet_info}")
            st.write(f"Colonnes détectées: {', '.join(df.columns[:10])}...")
            
            # Normaliser les colonnes
            df_normalized = SinistresService.normalize_columns(df)
            
            # Debug: Afficher les branches trouvées avant mapping
            branches_avant = df_normalized["Branche"].unique().to_list()
            st.write(f"  → Branches trouvées ({len(branches_avant)}): {branches_avant[:15]}{'...' if len(branches_avant) > 15 else ''}")
            
            # Appliquer le mapping des branches
            df_normalized = df_normalized.with_columns(
                pl.col("Branche").replace_strict(
                    mapping_dict,
                    default=pl.col("Branche")
                ).alias("Branche")
            )
            
            # Debug: Afficher les branches après mapping
            branches_apres = df_normalized["Branche"].unique().to_list()
            st.write(f"  → Branches après mapping ({len(branches_apres)}): {branches_apres}")
            
            # Ajouter l'exercice
            df_normalized = df_normalized.with_columns(
                pl.lit(year).alias("Exercice")
            )
            
            all_data.append(df_normalized)
        
        # Concaténer tous les fichiers
        combined = pl.concat(all_data)
        
        # Agréger par Branche et Exercice
        result = combined.group_by(["Branche", "Exercice"]).agg([
            pl.sum("Total_Reglement").alias("Total_Reglement"),
            pl.sum("Cout_Total").alias("Cout_Total"),
            pl.len().cast(pl.Int64).alias("Nombre_Sinistres")
        ])
        
        # Calculer SAP
        result = result.with_columns(
            (pl.col("Cout_Total") - pl.col("Total_Reglement")).alias("SAP")
        )
        
        return result.sort(["Branche", "Exercice"])
    
    @staticmethod
    def pivot_sinistres(df: pl.DataFrame, value_col: str) -> pl.DataFrame:
        """
        Pivote les données sinistres par branche et exercice.
        
        Args:
            df: DataFrame avec Branche, Exercice, et colonnes de valeurs
            value_col: Colonne à pivoter (Total_Reglement, Cout_Total, SAP)
            
        Returns:
            DataFrame pivoté avec ligne et colonne Total
        """
        pivot_df = df.pivot(
            values=value_col,
            index="Branche",
            on="Exercice"
        ).fill_null(0).sort("Branche")
        
        # Trier les colonnes exercices
        exercice_cols = sorted([col for col in pivot_df.columns if col != "Branche"])
        
        # Ajouter colonne Total
        pivot_df = pivot_df.with_columns(
            pl.sum_horizontal(exercice_cols).alias("Total")
        )
        
        # Réordonner les colonnes
        pivot_df = pivot_df.select(["Branche"] + exercice_cols + ["Total"])
        
        # Ajouter ligne Total
        totals = {"Branche": "TOTAL"}
        for col in exercice_cols + ["Total"]:
            totals[col] = pivot_df[col].sum()
        
        total_row = pl.DataFrame([totals])
        pivot_df = pl.concat([pivot_df, total_row])
        
        return pivot_df
